"""
目的：使用命令一键生成model模型文档

使用：1、注册app后运行命令  python manage.py model_to_doc appname1 appname2 
      2、根据提示选择生成的文档格式，文档默认存在项目根目录下的docs文件夹

注意：不加appname默认生成包含django自带的所有模型文档，尽量指定app使用
"""

import os
import sys
import time
from types import TracebackType
import openpyxl as opl
from django import setup
from django.apps import apps
from django.core.management.base import BaseCommand
from openpyxl.utils import get_column_letter

from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
)

# 设置单元格样式
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
)
fill_index = PatternFill("solid", fgColor="FFD700")
fill_table = PatternFill("solid", fgColor="228B22")
font_index = Font(name="微软雅黑", size=12, bold=True)
font_table = Font(name="微软雅黑", size=13, bold=True, italic=True)


def magic_doc(need_apps):
    """
    生成md.
    """
    models = apps.all_models
    if need_apps:
        model_dict = {
            x: list(models.get(x).values()) for x in models if x in need_apps
        }
    else:
        model_dict = {x: list(models.get(x).values()) for x in models}
    files_text = list()
    project_name = "Repair"
    files_text.append("# %s数据库设计" % (project_name,))
    files_text.append("- 文档时间： %s" % (time.strftime("%Y-%m-%d"),))

    files_text.append("## 引言")
    files_text.append("### 关于 ")
    files_text.append("  此文档主要介绍%s数据库定义。" % (project_name,))
    files_text.append("### 目标读者 ")
    files_text.append("  此文档提供给软件开发人员和系统维护人员使用。")
    files_text.append("### 术语定义")
    files_text.append("### 参考资料")
    files_text.append("## 数据库设计")

    # 遍历所有指定app下的model
    for app, models in model_dict.items():
        files_text.append("### %s" % (app,))
        # 遍历所有字段
        for target_cls in models:
            # 获取verbose_name
            model_key = "%s" % (target_cls._meta.db_table,)
            fields = dict()
            for field in target_cls._meta.fields:
                # 外键字段名加_id
                if type(field).__name__ == "ForeignKey":
                    f_name = field.name + "_id"
                else:
                    f_name = field.name
                if f_name not in fields.keys():
                    fields[f_name] = dict()
                fields[f_name].update(field.__dict__)
                fields[f_name]["field_type"] = str(type(field).__name__)

            files_text.append("#### %s" % (model_key,))
            files_text.append("- 表名： %s" % (target_cls._meta.db_table,))
            files_text.append("- 中文名： %s" % (target_cls._meta.verbose_name,))
            files_text.append("\n")
            files_text.append("|字段|中文|字段类型|空|默认值|主/外键|枚举值|字符最大长度|备注|")
            files_text.append("|-|-|-|:-:|:-:|:-:|-|-|-|")
            for (k, v) in fields.items():
                is_main_key = is_for_key = False
                # 获取字段默认值
                if "NOT_PROVIDED" in str(v["default"]):
                    v["default"] = ""
                # 获取枚举值
                if v["choices"] == None:
                    v["choices"] = ""
                if v["primary_key"] is True:
                    is_main_key = True
                if v["field_type"] == "ForeignKey":
                    is_for_key = True
                key_types = list()
                # 获取字段类型  主键/外键
                if is_main_key:
                    key_types.append("主键")
                if is_for_key:
                    key_types.append("外键")
                v["primary_key"] = ",".join(key_types)
                args = list()
                for tag in [
                    "name",
                    "verbose_name",
                    "field_type",
                    "null",
                    "default",
                    "primary_key",
                    "choices",
                    "max_length",
                    "help_text"
                ]:
                    args.append("") if tag == "max_length" and v[tag] == None else args.append(str(v[tag]))
                data = "|%s|" % "|".join(args)
                files_text.append(data)
            files_text.append("\n\n")
    return files_text


def auto_width(ws):
    """
    excel 单元格宽度自适应

    """

    lks = []
    for i in range(1, ws.max_column + 1):
        lk = 1
        for j in range(1, ws.max_row + 1):
            ws.cell(j, i).border = border
            ws.cell(j, i).alignment = Alignment(horizontal="center")
            sz = ws.cell(row=j, column=i).value
            if isinstance(sz, str):
                lk1 = len(sz.encode("gbk"))
            else:
                lk1 = len(str(sz))
            if lk < lk1:
                lk = lk1
        lks.append(lk)

    for i in range(1, ws.max_column + 1):
        k = get_column_letter(i)
        ws.column_dimensions[k].width = lks[i - 1] + 2
    return ws


def to_excel(need_apps):
    """
    生成excel.
    """
    models = apps.all_models
    wb = opl.Workbook()
    if need_apps:
        model_dict = {
            x: list(models.get(x).values()) for x in models if x in need_apps
        }
    else:
        model_dict = {x: list(models.get(x).values()) for x in models}
    for app, models in model_dict.items():
        ws = wb.create_sheet(app, index=0)
        for target_cls in models:
            table = target_cls._meta.db_table
            table_verbose_name = target_cls._meta.verbose_name
            ws.append([table, str(table_verbose_name)])
            for column in [1, 2]:
                ws.cell(ws.max_row, column).alignment = Alignment(horizontal="center")
                ws.cell(ws.max_row, column).fill = fill_table
                ws.cell(ws.max_row, column).font = font_table
            ws.append(["字段", "中文", "字段类型", "空", "默认值", "主/外键", "枚举值", "字符最大长度", "备注"])
            for c in ws[ws.max_row]:
                c.alignment = Alignment(horizontal="center")
                c.fill = fill_index
                c.font = font_index
            fields = dict()
            for field in target_cls._meta.fields:
                if type(field).__name__ == "ForeignKey":
                    f_name = field.name + "_id"
                else:
                    f_name = field.name
                if f_name not in fields.keys():
                    fields[f_name] = dict()
                fields[f_name].update(field.__dict__)
                fields[f_name]["field_type"] = str(type(field).__name__)
                print(field.db_type())
            for (k, v) in fields.items():
                is_main_key = is_for_key = False
                if "NOT_PROVIDED" in str(v["default"]):
                    v["default"] = ""
                if v["choices"] == None:
                    v["choices"] = ""
                if v["primary_key"] is True:
                    is_main_key = True
                if v["field_type"] == "ForeignKey":
                    is_for_key = True
                key_types = list()
                if is_main_key:
                    key_types.append("主键")
                if is_for_key:
                    key_types.append("外键")
                v["primary_key"] = ",".join(key_types)
                args = list()
                for tag in [
                    "name",
                    "verbose_name",
                    "field_type",
                    "null",
                    "default",
                    "primary_key",
                    "choices",
                    "max_length",
                    "help_text"
                ]:
                    args.append("") if tag == "max_length" and v[tag] == None else args.append(str(v[tag]))
                ws.append(args)
            ws.append(["\n"])
            ws.append(["\n"])
            ws = auto_width(ws)
    wb.save("./docs/magic_model_excel.xlsx")


class Command(BaseCommand):
    def add_arguments(self, parser):

        parser.add_argument(
            dest="apps",  # 参数名字
            type=str,  # 参数类型
            help="需要转换文档的app",  # 帮助信息
            nargs="*",  # 参数列表
            default="",  # 默认值
        )

    def handle(self, *args, **options):
        app_name = " ".join([i for i in options["apps"]])
        if app_name:
            self.stdout.write(
                self.style.SUCCESS("需要转换model的app为：%s \n" % app_name)
            )
        else:
            self.stdout.write(self.style.SUCCESS("默认转换所有app的model"))
        need_apps = options["apps"]
        self.stdout.write(
            self.style.MIGRATE_HEADING("选择文档格式:  (1) md   (2) excel")
        )
        choice = sys.stdin.readline().replace("\n", "")
        try:
            if choice == "1" or choice == "md":
                txt = magic_doc(need_apps)
                with open(
                        "./docs/magic_model_markdown.md", "w", encoding="utf-8"
                ) as fw:
                    fw.write("\n".join(txt))
                self.stdout.write(self.style.SUCCESS("转换成功"))
            elif choice == "2" or choice == "excel":
                to_excel(need_apps)
                self.stdout.write(self.style.SUCCESS("转换成功"))
            else:
                self.stderr.write(self.style.ERROR("不支持的格式"))
        except Exception as e:
            print(e)
